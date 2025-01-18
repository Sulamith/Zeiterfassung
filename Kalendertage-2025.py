from openpyxl import load_workbook
import calendar

def aktualisiere_kalender(datei_pfad, jahr):
    # Arbeitsmappe laden
    wb = load_workbook(datei_pfad)

    # Durchlaufe alle Blätter der Monate Januar bis Dezember
    for monat in calendar.month_name[1:]:
        if monat in wb.sheetnames:  # Nur wenn das Monatsblatt existiert
            ws = wb[monat]

            # Kalendertage für den jeweiligen Monat und das Jahr
            _, letzte_tag_nummer = calendar.monthrange(jahr, list(calendar.month_name).index(monat))
            tage = list(range(1, letzte_tag_nummer + 1))

            # Kopiere die Liste der zusammengeführten Bereiche, bevor sie geändert wird
            merged_ranges = list(ws.merged_cells.ranges)
            for merge in merged_ranges:
                if merge.min_col <= 2 and merge.max_col >= 2:  # Überprüfe Spalten A und B
                    ws.unmerge_cells(str(merge))

            # Spalten 'Tag' und 'Datum' aktualisieren
            for zeile, tag in enumerate(tage, start=5):  # Startzeile für die Daten ist Zeile 5
                ws[f"A{zeile}"] = calendar.day_abbr[calendar.weekday(jahr, list(calendar.month_name).index(monat), tag)]  # Wochentag
                ws[f"B{zeile}"] = tag  # Tag

            # Entferne überschüssige Einträge aus vorherigen Jahren
            for zeile in range(len(tage) + 5, ws.max_row + 1):
                ws[f"A{zeile}"] = None
                ws[f"B{zeile}"] = None

    # Arbeitsmappe speichern
    wb.save(datei_pfad)
    print(f"Die Tabelle wurde erfolgreich aktualisiert: {datei_pfad}")

# Datei und Jahr anpassen
datei_pfad = "/Users/l-gehsul00/Documents/GitHub_Sulamith/Administratives/Stundenabrechnung_Vorlage_2025__BEBB_aktualisiert.xlsx"
jahr = 2025

aktualisiere_kalender(datei_pfad, jahr)
